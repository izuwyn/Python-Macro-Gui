import tkinter as tk
from tkinter import filedialog, scrolledtext
import threading
import time
import pyautogui
import pyperclip
import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime
import keyboard

# ── Default Config ────────────────────────────────────────────────────────────
DEFAULT_PACK_VALUE   = 'NVC PACK'
DEFAULT_MAX_RECORDS  = 1
HOTKEY_START         = 'f3'
HOTKEY_STOP          = 'f4'
HOTKEY_CLEAR         = 'f5'

T_CLICK    = 0.15
T_KEY      = 0.05
T_SEARCH   = 0.20
T_NAVIGATE = 0.30

COORD_EXCEL_CELL  = (-259, -147)
COORD_ACCESS_MAIN = (104, 212)
COORD_ACCESS_WIN  = (1219, 1050)
COORD_EXCEL_WIN   = (-251, 465)

PIXEL_SEARCH_OPEN = {'pos': (47, 351),  'color': (38, 38, 38)}
PIXEL_RECORD_OPEN = {'pos': (128, 210), 'color': (38, 38, 38)}
# ─────────────────────────────────────────────────────────────────────────────

class MacroApp:
    def __init__(self, root):
        self.root = root
        self.root.title("AutoPack Runner")
        self.root.geometry("560x700")
        self.root.resizable(False, False)
        self.root.configure(bg="#0f0f0f")

        self.running = False
        self.stop_flag = threading.Event()
        self.excel_path = tk.StringVar(value="")
        self.pack_value = tk.StringVar(value=DEFAULT_PACK_VALUE)
        self.max_records = tk.IntVar(value=DEFAULT_MAX_RECORDS)
        self.start_record = tk.IntVar(value=1)
        
        self.processed = tk.IntVar(value=0)
        self.skipped = tk.IntVar(value=0)
        self.errors = tk.IntVar(value=0)

        self._build_ui()
        self._register_hotkeys()

    # ── UI ────────────────────────────────────────────────────────────────────

    def _build_ui(self):
        C = {
            'bg':       '#0f0f0f',
            'panel':    '#1a1a1a',
            'border':   '#2a2a2a',
            'accent':   '#00c896',
            'danger':   '#ff4f4f',
            'text':     '#e8e8e8',
            'muted':    '#666666',
            'input_bg': '#111111',
        }
        self.C = C

        hdr = tk.Frame(self.root, bg="#0f0f0f")
        hdr.pack(fill='x', padx=20, pady=(20, 0))

        tk.Label(hdr, text="AUTOPACK", font=("Courier New", 22, "bold"),
                 fg=C['accent'], bg=C['bg']).pack(side='left')
        tk.Label(hdr, text="RUNNER", font=("Courier New", 22),
                 fg=C['text'], bg=C['bg']).pack(side='left', padx=(6, 0))

        hotkey_lbl = tk.Label(hdr,
            text=f"{HOTKEY_START.upper()} start · {HOTKEY_STOP.upper()} stop · {HOTKEY_CLEAR.upper()} reset",
            font=("Courier New", 8), fg=C['muted'], bg=C['bg'])
        hotkey_lbl.pack(side='right', pady=(8, 0))

        tk.Frame(self.root, bg=C['border'], height=1).pack(fill='x', padx=20, pady=12)

        panel = tk.Frame(self.root, bg=C['panel'], bd=0, relief='flat')
        panel.pack(fill='x', padx=20)

        self._section_label(panel, "EXCEL FILE")
        file_row = tk.Frame(panel, bg=C['panel'])
        file_row.pack(fill='x', padx=12, pady=(0, 10))
        self.file_entry = tk.Entry(file_row, textvariable=self.excel_path, bg=C['input_bg'], fg=C['text'], relief='flat', font=("Courier New", 9), bd=0)
        self.file_entry.pack(side='left', fill='x', expand=True, ipady=6, ipadx=6)
        tk.Button(file_row, text="Browse", command=self._browse, bg=C['border'], fg=C['text'], relief='flat', font=("Courier New", 9), cursor='hand2', padx=10).pack(side='left', padx=(6, 0))

        conf_row = tk.Frame(panel, bg=C['panel'])
        conf_row.pack(fill='x', padx=12, pady=(0, 10))

        left_col = tk.Frame(conf_row, bg=C['panel'])
        left_col.pack(side='left', fill='x', expand=True, padx=(0, 5))
        self._section_label(left_col, "START AT RECORD", pad=0)
        tk.Spinbox(left_col, from_=1, to=99999, textvariable=self.start_record, bg=C['input_bg'], fg=C['accent'], relief='flat', font=("Courier New", 10), bd=0).pack(fill='x', ipady=6)

        right_col = tk.Frame(conf_row, bg=C['panel'])
        right_col.pack(side='left', fill='x', expand=True, padx=(5, 0))
        self._section_label(right_col, "MAX TO PROCESS", pad=0)
        tk.Spinbox(right_col, from_=1, to=10000, textvariable=self.max_records, bg=C['input_bg'], fg=C['text'], relief='flat', font=("Courier New", 10), bd=0).pack(fill='x', ipady=6)

        self._section_label(panel, "PACK VALUE")
        tk.Entry(panel, textvariable=self.pack_value, bg=C['input_bg'], fg=C['accent'], relief='flat', font=("Courier New", 10, "bold"), bd=0).pack(fill='x', padx=12, ipady=6, ipadx=6, pady=(0, 14))

        stats = tk.Frame(self.root, bg=C['bg'])
        stats.pack(fill='x', padx=20, pady=10)
        self._stat_box(stats, "PROCESSED", self.processed, C['accent'])
        self._stat_box(stats, "SKIPPED",   self.skipped,   '#f0a500')
        self._stat_box(stats, "ERRORS",    self.errors,    C['danger'])

        self.prog_canvas = tk.Canvas(self.root, bg=C['border'], height=4, bd=0, highlightthickness=0)
        self.prog_canvas.pack(fill='x', padx=20, pady=(0, 10))
        self.prog_bar = self.prog_canvas.create_rectangle(0, 0, 0, 4, fill=C['accent'], width=0)

        self.log = scrolledtext.ScrolledText(self.root, bg=C['input_bg'], fg=C['text'], font=("Courier New", 9), relief='flat', state='disabled', wrap='word')
        self.log.pack(fill='both', expand=True, padx=20, pady=(0, 10))
        self.log.tag_config('ok', foreground=C['accent'])
        self.log.tag_config('skip', foreground='#f0a500')
        self.log.tag_config('err', foreground=C['danger'])
        self.log.tag_config('head', foreground=C['text'])

        btn_row = tk.Frame(self.root, bg=C['bg'])
        btn_row.pack(fill='x', padx=20, pady=(0, 20))
        self.start_btn = tk.Button(btn_row, text="▶  START", command=self.start_macro, bg=C['accent'], font=("Courier New", 11, "bold"), pady=10)
        self.start_btn.pack(side='left', fill='x', expand=True, padx=(0, 6))
        self.stop_btn = tk.Button(btn_row, text="■  STOP", command=self.stop_macro, bg=C['border'], font=("Courier New", 11, "bold"), pady=10, state='disabled')
        self.stop_btn.pack(side='left', fill='x', expand=True)

    def _section_label(self, parent, text, pad=12):
        tk.Label(parent, text=text, font=("Courier New", 8, "bold"), fg=self.C['muted'], bg=self.C['panel']).pack(anchor='w', padx=pad, pady=(10, 2))

    def _stat_box(self, parent, label, var, color):
        box = tk.Frame(parent, bg=self.C['panel'], padx=14, pady=8)
        box.pack(side='left', fill='x', expand=True, padx=(0, 6))
        tk.Label(box, textvariable=var, font=("Courier New", 20, "bold"), fg=color, bg=self.C['panel']).pack()
        tk.Label(box, text=label, font=("Courier New", 7), fg=self.C['muted'], bg=self.C['panel']).pack()

    # ── Logic ─────────────────────────────────────────────────────────────────

    def _browse(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xlsm")])
        if path: self.excel_path.set(path)

    def reset_progress(self):
        if self.running: return
        self.processed.set(0)
        self.skipped.set(0)
        self.errors.set(0)
        self.start_record.set(1)
        self._set_progress(0, 1)
        self._log("↺ Reset to record #1", 'head')

    def _log(self, msg, tag='info'):
        def _write():
            self.log.configure(state='normal')
            self.log.insert('end', f"{msg}\n", tag)
            self.log.see('end')
            self.log.configure(state='disabled')
        self.root.after(0, _write)

    def _set_progress(self, done, total):
        def _draw():
            w = self.prog_canvas.winfo_width()
            pct = done / total if total else 0
            self.prog_canvas.coords(self.prog_bar, 0, 0, int(w * pct), 4)
        self.root.after(0, _draw)

    def _register_hotkeys(self):
        keyboard.add_hotkey(HOTKEY_START, self.start_macro)
        keyboard.add_hotkey(HOTKEY_STOP,  self.stop_macro)
        keyboard.add_hotkey(HOTKEY_CLEAR, self.reset_progress)

    def _set_running(self, state: bool):
        def _apply():
            if state:
                self.start_btn.configure(state='disabled', bg=self.C['border'])
                self.stop_btn.configure(state='normal', bg=self.C['danger'])
            else:
                self.start_btn.configure(state='normal', bg=self.C['accent'])
                self.stop_btn.configure(state='disabled', bg=self.C['border'])
        self.root.after(0, _apply)

    def start_macro(self):
        if self.running or not self.excel_path.get(): return
        self.running = True
        self.stop_flag.clear()
        self._set_running(True)
        threading.Thread(target=self._run, daemon=True).start()

    def stop_macro(self):
        self.stop_flag.set()
        self._log("⏹ Stop requested — finishing record...", 'err')

    def _run(self):
        pyautogui.FAILSAFE = True
        try:
            wb, ws, values = self._load_excel()
        except Exception as e:
            self._log(f"✗ Load Error: {e}", 'err')
            self._set_running(False); self.running = False
            return

        total = len(values)
        self._log(f"── Starting at Record #{self.start_record.get()} ──", 'head')

        done_count = self.processed.get()
        skip_count = self.skipped.get()
        err_count  = self.errors.get()

        for i, (row_num, value, col_b, col_c) in enumerate(values):
            if self.stop_flag.is_set(): break

            current_idx = self.start_record.get()
            self._log(f"[{current_idx}] A={value} | B={col_b} | C={col_c}", 'head')
            
            result = self._process_record(value, col_b, col_c)

            if result == 'done':
                done_count += 1
            elif result == 'skip':
                skip_count += 1
                self._color_cell(ws, row_num, "FF0000")
            else:
                err_count += 1
                self._color_cell(ws, row_num, "FF4F4F")
                self._log(f"  ✗ {result}", 'err')

            self.root.after(0, lambda: self.start_record.set(current_idx + 1))
            self.root.after(0, lambda d=done_count: self.processed.set(d))
            self.root.after(0, lambda s=skip_count: self.skipped.set(s))
            self.root.after(0, lambda e=err_count: self.errors.set(e))
            self._set_progress(i + 1, total)

        wb.save(self.excel_path.get())
        self._log("── Session Ended / Excel Saved ──", 'head')
        self._set_running(False); self.running = False

    def _format_cell(self, val):
        """Convert cell value to a clean string, formatting dates as M/D."""
        if val is None:
            return ""
        if isinstance(val, datetime):
            return val.strftime("%#m/%#d")  # Windows: no leading zeros → "4/17"
        return str(val).strip()

    def _load_excel(self):
        wb = openpyxl.load_workbook(self.excel_path.get())
        ws = wb.active
        values = []

        start_row = self.start_record.get()

        for i, row in enumerate(ws.iter_rows(min_row=start_row), start=start_row):
            if row[0].value:
                col_a = row[0].value
                col_b = self._format_cell(row[1].value if len(row) > 1 else None)
                col_c = self._format_cell(row[2].value if len(row) > 2 else None)
                values.append((i, col_a, col_b, col_c))
            if len(values) >= self.max_records.get():
                break

        return wb, ws, values

    def _color_cell(self, ws, row_num, hex_color):
        ws.cell(row=row_num, column=1).fill = PatternFill("solid", fgColor=hex_color)

    def _process_record(self, value, col_b, col_c):
        try:
            pyperclip.copy("")
            pyperclip.copy(str(value))

            self._click(COORD_ACCESS_MAIN, delay=0.5)

            self._hotkey('ctrl', 'f', delay=0.7)

            search_ready = False
            for _ in range(5):
                if self._pixel_matches(PIXEL_SEARCH_OPEN):
                    search_ready = True
                    break
                time.sleep(0.3)

            if not search_ready:
                return 'Search Dialog Timed Out'

            self._hotkey('ctrl', 'a', delay=0.2)
            self._hotkey('ctrl', 'v', delay=0.2)
            self._press('enter', delay=0.8)

            record_found = False
            for _ in range(5):
                if self._pixel_matches(PIXEL_RECORD_OPEN):
                    record_found = True
                    break
                time.sleep(0.5)

            if not record_found:
                self._press('escape', delay=0.5)
                return 'not_found'

            self._press('escape', delay=0.5)

            for _ in range(6):
                self._press('tab', delay=0.01)

            pyperclip.copy("")
            self._hotkey('ctrl', 'c', delay=0.2)

            clipboard_content = ""
            for _ in range(5):
                clipboard_content = pyperclip.paste().strip()
                if clipboard_content:
                    break
                time.sleep(0.1)

            if "TRIAGE COMP" in clipboard_content.upper():
                self._log(f"  → Found '{clipboard_content}', skipping.", 'skip')
                return 'skip'

            # Write PACK value
            pyperclip.copy(self.pack_value.get())
            self._hotkey('ctrl', 'v', delay=0.05)
            self._press('tab', delay=0.2)        # commits PACK, lands on next field

            # ── Column B ──────────────────────────────────────────────────────
            if col_b:
                for _ in range(3):
                    self._press('tab', delay=0.15)
                pyperclip.copy(col_b)
                self._hotkey('ctrl', 'v', delay=0.1)

            # ── Column C ──────────────────────────────────────────────────────
            if col_c:
                for _ in range(2):
                    self._press('tab', delay=0.15)
                pyperclip.copy(col_c)
                self._hotkey('ctrl', 'v', delay=0.1)

            return 'done'

        except Exception as e:
            return f"Error: {str(e)}"

    # Basic UI Action Wrappers
    def _click(self, coord, delay=T_CLICK):
        pyautogui.click(coord[0], coord[1]); time.sleep(delay)
    def _hotkey(self, *keys, delay=T_KEY):
        pyautogui.hotkey(*keys); time.sleep(delay)
    def _press(self, key, delay=T_KEY):
        pyautogui.press(key); time.sleep(delay)
    def _pixel_matches(self, check):
        return pyautogui.pixel(check['pos'][0], check['pos'][1]) == check['color']

if __name__ == '__main__':
    root = tk.Tk()
    app = MacroApp(root)
    root.mainloop()